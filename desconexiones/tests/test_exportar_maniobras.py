from io import BytesIO

from django.test import SimpleTestCase
from django.urls import reverse
from openpyxl import load_workbook


class ExportarManiobrasExcelTests(SimpleTestCase):
    def test_exporta_toda_la_informacion_disponible(self):
        fila = [1, "27/08/2026", "PATIOS", "PATC12", "RC-101", "Reconectador", "ABRIR", 13.8, 12345, "Calle 1", "Abrir reconectador", "Aislar zona"]
        respuesta = self.client.post(
            reverse("api_exportar_maniobras_xlsx"),
            data={"nombre": "prueba", "secciones": [{"titulo": "Desenergización", "filas": [fila]}]},
            content_type="application/json",
        )
        self.assertEqual(respuesta.status_code, 200)
        libro = load_workbook(BytesIO(respuesta.content), read_only=True)
        hoja = libro["Desenergización"]
        self.assertEqual(
            list(next(hoja.iter_rows(values_only=True))),
            ["#", "Fecha", "Subestación", "Celda", "Dispositivo", "Tipo", "Estado", "Nivel kV", "FID GTECH", "Dirección aproximada", "Maniobra", "Motivo técnico"],
        )
        self.assertEqual(list(next(hoja.iter_rows(min_row=2, values_only=True))), fila)
