from io import BytesIO

from django.test import SimpleTestCase
from django.urls import reverse
from openpyxl import load_workbook


class ExportarUsuariosExcelTests(SimpleTestCase):
    def test_exporta_usuarios_desconectados_sin_duplicar_niu(self):
        usuarios = [
            {"numero_transformador":"7356","nodo_transformador":"1T01591","fases":"RST","circuito":"PELC1","niu":"123","razon_social":"USUARIO UNO","municipio":"Pelaya"},
            {"numero_transformador":"7356","nodo_transformador":"1T01591","fases":"RST","circuito":"PELC1","niu":"123","razon_social":"USUARIO UNO","municipio":"Pelaya"},
            {"numero_transformador":"8000","nodo_transformador":"1T09999","fases":"ST","circuito":"PELC2","niu":"456","razon_social":"USUARIO DOS","municipio":""},
        ]
        respuesta = self.client.post(
            reverse("api_exportar_usuarios_xlsx"),
            data={"nombre": "usuarios-desconectados", "usuarios": usuarios},
            content_type="application/json",
        )

        self.assertEqual(respuesta.status_code, 200)
        libro = load_workbook(BytesIO(respuesta.content), read_only=True)
        filas = list(libro.active.iter_rows(values_only=True))
        self.assertEqual(len(filas[0]),30)
        self.assertEqual(filas[0][0],"Numero de transformador")
        self.assertEqual(filas[0][4],"Feha/Hora Desenergizacón")
        self.assertEqual(len(filas), 3)
        self.assertEqual(filas[1][1], "1T01591")
        self.assertIsNone(filas[1][4])
        self.assertEqual(filas[1][13],"123")
