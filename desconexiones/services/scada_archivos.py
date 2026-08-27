from __future__ import annotations

import io
import re
import statistics
import unicodedata
import zipfile
from collections import defaultdict
from datetime import date, datetime, time
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .configuracion_local import estado_carpeta_scada


VARIABLES = {"URS", "UST", "UTR", "IR", "IS", "IT", "P", "Q", "S", "FP"}
EQUIVALENCIAS = {
    "URS": "URS", "UAB": "URS", "UR": "URS", "VRS": "URS",
    "UST": "UST", "UBC": "UST", "US": "UST", "UB": "UST", "VST": "UST",
    "UTR": "UTR", "UCA": "UTR", "UT": "UTR", "UC": "UTR", "VTR": "UTR",
    "IR": "IR", "IS": "IS", "IT": "IT", "P": "P", "Q": "Q", "S": "S",
    "FP": "FP", "PF": "FP",
}
FUENTES = {"REL", "MED", "REC", "IED", "RTU", "PM"}
PRIORIDAD = ("MED", "REL", "REC", "IED", "RTU", "PM", "")
MESES = {
    1: ("ENERO", "ENE"), 2: ("FEBRERO", "FEB"), 3: ("MARZO", "MAR"),
    4: ("ABRIL", "ABR"), 5: ("MAYO", "MAY"), 6: ("JUNIO", "JUN"),
    7: ("JULIO", "JUL"), 8: ("AGOSTO", "AGO"),
    9: ("SEPTIEMBRE", "SEP", "SEPT"), 10: ("OCTUBRE", "OCT"),
    11: ("NOVIEMBRE", "NOV"), 12: ("DICIEMBRE", "DIC"),
}
ALIAS = {
    ("SANROQUE", "SANRAUX"): "SRQAUX", ("SANROQUE", "SANRC1"): "SRQC1",
    ("SANROQUE", "SANRC2"): "SRQC2", ("SANROQUE", "SANRC3"): "SRQC3",
    ("SANROQUE", "SANRIL10"): "SRQIL10", ("SANROQUE", "SANRIL30"): "SRQIL30",
    ("SANROQUE", "SANRIT20"): "SRQIT20", ("SANROQUE", "SANRPPAL"): "SRQPPAL",
    ("SEVILLA", "SEVC30"): "SEVC20", ("SEVILLA", "SEVC31"): "SEVC21",
    ("SEVILLA", "SEVC32"): "SEVC22", ("PLANTAZULIA", "PLZ306"): "PLZ306B1",
    ("LOSPATIOS", "PATIOS345"): "PATIT40",
}


def normalizar(valor: Any) -> str:
    texto = unicodedata.normalize("NFKD", str(valor or "").strip().upper())
    return "".join(c for c in texto if not unicodedata.combining(c))


def clave(valor: Any) -> str:
    return re.sub(r"[^A-Z0-9]", "", normalizar(valor))


def claves_dispositivo(valor: Any) -> set[str]:
    base = clave(valor)
    salida = {base}
    if base.startswith("CSERVAUX"):
        salida.add(base[1:])
    return salida


def numero(valor: Any) -> float | None:
    if valor in (None, ""):
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip().replace(" ", "")
    if "," in texto and "." not in texto:
        texto = texto.replace(",", ".")
    elif "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return None


def convertir_datetime(valor: Any) -> datetime | None:
    if valor in (None, ""):
        return None
    if isinstance(valor, datetime):
        return valor
    if isinstance(valor, date):
        return datetime.combine(valor, time.min)
    texto = str(valor).strip()
    for formato in (
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M", "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M",
        "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M", "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M", "%Y-%m-%dT%H:%M:%S",
    ):
        try:
            return datetime.strptime(texto, formato)
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(texto)
    except ValueError:
        return None


def interpretar_senal(encabezado: Any) -> tuple[str, str, str] | None:
    texto = re.sub(r"\s+", "_", normalizar(encabezado))
    if "_" not in texto:
        return None
    partes = texto.split("_")
    variable = EQUIVALENCIAS.get(partes[0])
    if not variable:
        return None
    if len(partes) >= 3 and partes[1] in FUENTES:
        fuente, dispositivo = partes[1], "_".join(partes[2:])
    else:
        fuente, dispositivo = "", "_".join(partes[1:])
    return variable, fuente, dispositivo.strip("_")


def detectar_columna_fecha(encabezados: list[Any]) -> int | None:
    exactos = {"TIME", "TIMESTAMP", "DATE", "DATETIME", "FECHA", "FECHAHORA", "FECHAYHORA", "HORA"}
    for i, encabezado in enumerate(encabezados):
        if clave(encabezado) in exactos:
            return i
    for i, encabezado in enumerate(encabezados):
        texto = normalizar(encabezado)
        if "FECHA" in texto or "TIME" in texto or "DATE" in texto:
            return i
    return None


def carpeta_mes(nombre: str, anio: int, mes: int) -> bool:
    texto = normalizar(nombre)
    if "SCADA" not in texto:
        return False
    numero_mes = re.match(r"^0?(\d{1,2})[. _-]", texto)
    if numero_mes and int(numero_mes.group(1)) != mes:
        return False
    anios = re.findall(r"20\d{2}", texto)
    if anios and str(anio) not in anios:
        return False
    return bool((numero_mes and int(numero_mes.group(1)) == mes) or any(x in texto for x in MESES[mes]))


def archivos_fecha(raiz: Path, fecha: date) -> list[Path]:
    carpeta_anio = raiz / str(fecha.year)
    if not carpeta_anio.is_dir():
        raise FileNotFoundError(f"No existe la carpeta SCADA del año {fecha.year}: {carpeta_anio}")
    meses = [p for p in carpeta_anio.iterdir() if p.is_dir() and carpeta_mes(p.name, fecha.year, fecha.month)]
    if not meses:
        raise FileNotFoundError(f"No se encontró la carpeta SCADA de {fecha:%Y-%m} dentro de {carpeta_anio}.")
    patron = re.compile(r"^\d{1,3}_.+\.(xlsx|xlsm)$", re.IGNORECASE)
    return sorted({p for mes in meses for p in mes.rglob("*") if p.is_file() and p.suffix.lower() in {".xlsx", ".xlsm"} and not p.name.startswith("~$") and patron.match(p.name)})


_PATRON_V = re.compile(rb"(<(?:[A-Za-z_][\w.-]*:)?v(?:\s+[^>]*)?>)(.*?)(</(?:[A-Za-z_][\w.-]*:)?v>)", re.I | re.S)


def abrir_excel(path: Path):
    try:
        return load_workbook(path, read_only=True, data_only=True)
    except (ValueError, TypeError):
        memoria = io.BytesIO()
        with zipfile.ZipFile(path, "r") as entrada, zipfile.ZipFile(memoria, "w", zipfile.ZIP_DEFLATED) as salida:
            for info in entrada.infolist():
                datos = entrada.read(info.filename)
                if info.filename.lower().startswith("xl/worksheets/") and info.filename.lower().endswith(".xml"):
                    datos = _PATRON_V.sub(lambda m: m.group(1) + re.sub(rb"(?<=\d),(?=\d)", b".", m.group(2)) + m.group(3), datos)
                salida.writestr(info, datos)
        memoria.seek(0)
        libro = load_workbook(memoria, read_only=True, data_only=True)
        libro._cdym_memoria = memoria
        return libro


def subestacion_archivo(nombre: str) -> str:
    texto = Path(nombre).stem.upper()
    texto = re.sub(r"^\d{1,3}_", "", texto)
    texto = re.sub(r"20\d{12}$", "", texto)
    texto = re.sub(r"[_\-\s]*(?:13|34|13[,.]8|34[,.]5|69|115|138|230)\s*(?:KV)?$", "", texto, flags=re.I)
    texto = re.sub(r"[_\-\s]+(?:13[,.]?8|34[,.]?5|69|115|138|230)\s*KV.*$", "", texto, flags=re.I)
    texto = re.sub(r"[_\-\s]+MODULO.*$", "", texto, flags=re.I)
    return clave(texto)


def dispositivo_equivalente(subestacion: str, encontrado: str, solicitado: str) -> bool:
    sub = clave(subestacion)
    encontrado_clave = clave(encontrado)
    encontrado_clave = ALIAS.get((sub, encontrado_clave), encontrado_clave)
    return bool(claves_dispositivo(encontrado_clave) & claves_dispositivo(solicitado))


@lru_cache(maxsize=256)
def leer_dia(raiz_texto: str, fecha_iso: str, subestacion: str, dispositivo: str, fuente_preferida: str, nivel_kv: float | None) -> dict:
    raiz, fecha = Path(raiz_texto), date.fromisoformat(fecha_iso)
    por_fuente: dict[str, dict[datetime, dict[str, float]]] = defaultdict(lambda: defaultdict(dict))
    origenes: dict[str, set[str]] = defaultdict(set)
    archivos = archivos_fecha(raiz, fecha)
    sub_clave = clave(subestacion)
    candidatos = [p for p in archivos if subestacion_archivo(p.name) == sub_clave]
    if not candidatos:
        candidatos = archivos
    errores = []
    for archivo in candidatos:
        libro = None
        try:
            libro = abrir_excel(archivo)
            for hoja in libro.worksheets:
                encabezados = [c.value for c in next(hoja.iter_rows(min_row=1, max_row=1))]
                indice_fecha = detectar_columna_fecha(encabezados)
                if indice_fecha is None:
                    continue
                columnas = {}
                for indice, encabezado in enumerate(encabezados):
                    senal = interpretar_senal(encabezado)
                    if senal and dispositivo_equivalente(subestacion, senal[2], dispositivo):
                        columnas[indice] = senal
                if not columnas:
                    continue
                for fila in hoja.iter_rows(min_row=2, values_only=True):
                    if indice_fecha >= len(fila):
                        continue
                    instante = convertir_datetime(fila[indice_fecha])
                    if instante is None or instante.date() != fecha:
                        continue
                    for indice, (variable, fuente, _) in columnas.items():
                        if indice >= len(fila):
                            continue
                        valor = numero(fila[indice])
                        if valor is not None:
                            por_fuente[fuente][instante][variable] = valor
                            origenes[fuente].add(archivo.name)
        except Exception as exc:
            errores.append(f"{archivo.name}: {exc}")
        finally:
            if libro is not None:
                try:
                    libro.close()
                except Exception:
                    pass
    completas = {f: puntos for f, puntos in por_fuente.items() if len(puntos) == 96}
    prioridades = []
    for fuente in (fuente_preferida.strip().upper(), *PRIORIDAD):
        if fuente not in prioridades:
            prioridades.append(fuente)
    fuente = next((f for f in prioridades if f in completas), None)
    if fuente is None:
        resumen_fuentes = ", ".join(f"{f or 'SIN_FUENTE'}={len(p)}" for f, p in por_fuente.items())
        detalle = f" Fuentes encontradas: {resumen_fuentes}." if resumen_fuentes else ""
        raise ValueError(f"{dispositivo}: no se encontraron 96 intervalos completos para {fecha_iso} en la carpeta SCADA.{detalle}")
    puntos = completas[fuente]
    return {
        "subestacion": subestacion, "dispositivo": dispositivo, "fuente": fuente,
        "nivel_kv": nivel_kv, "cantidad_registros": 96,
        "registros": [(instante.strftime("%Y-%m-%d %H:%M:%S"), valores) for instante, valores in sorted(puntos.items())],
        "archivos_origen": sorted(origenes[fuente]), "errores_archivos": errores,
    }


def serie_dia_scada(medida: dict, fecha: date, campos: tuple[tuple[str, str], ...], unidades: dict[str, str]) -> dict:
    estado = estado_carpeta_scada()
    if not estado["valida"]:
        raise FileNotFoundError(f"Configure en Admin > Datos SCADA una carpeta disponible. {estado['mensaje']}")
    subestacion = str(medida.get("medida_subestacion") or medida.get("subestacion") or "").strip().upper()
    dispositivo = str(medida.get("medida_dispositivo") or medida.get("interruptor") or "").strip().upper()
    fuente = str(medida.get("medida_fuente") or medida.get("fuente") or "").strip().upper()
    nivel = medida.get("nivel_kv")
    datos = leer_dia(estado["ruta"], fecha.isoformat(), subestacion, dispositivo, fuente, float(nivel) if nivel is not None else None)
    series = []
    for nombre, etiqueta in campos:
        puntos = [[instante, valores[nombre]] for instante, valores in datos["registros"] if nombre in valores]
        series.append({"clave": nombre, "nombre": etiqueta, "unidad": unidades[nombre], "puntos": puntos})
    return {**{k: v for k, v in datos.items() if k != "registros"}, "series": series, "origen_datos": "ARCHIVOS_SCADA"}
