import json
import os
import threading
from io import BytesIO
from datetime import datetime
from pathlib import Path
from django.conf import settings
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.views.decorators.http import require_GET, require_POST
from django.db import transaction
from django.utils import timezone
from .services.gtech import FNO_CORTE, GTechService
from .services.sac import SacService
from .services.medidas import cargar_catalogo, medidas_subestacion, coincidencia_preferida, normalizar
from .services.series_medidas import SeriesMedidasService, VARIABLES
from .services.sets_proteccion import set_sugerido
from .services.kmz import leer_kmz
from .models import AsignacionMedidaEnergia, RelacionBarraTransformacion, RamalTransformadorManual, GrupoTransformadorBarra, BarraGrupoTransformador, CatalogoConductorCens, ConfiguracionTendidoCircuito, EstadoOperativoVigente, ParaleloCeldaPermitido, ConfiguracionManiobras, AprendizajeProtocolo, EventoAprendizajeProtocolo, PerfilAprendizajeManiobras
from .services.ia.aprendizaje import reconstruir_preferencias

CIRCUITOS_DIR = Path(settings.DATA_DIR) / "circuitos"
_CIRCUITOS_CACHE_LOCK = threading.RLock()

def _archivo_circuito(subestacion, fid):
    codigo = "".join(c for c in subestacion.upper() if c.isalnum() or c in "-_")
    return CIRCUITOS_DIR / f"{codigo}__{int(fid)}.json"

def _completar_sac(data):
    transformadores=[e for e in data["elementos"] if e.get("g3e_fno")==20400]
    codigos=[e.get("codigo_operacion") or e.get("codigo") for e in transformadores]
    usuarios_por_trafo,excluidos_por_trafo,estado_por_trafo=SacService().usuarios_por_transformadores(codigos)
    for elemento in transformadores:
        codigo=str(elemento.get("codigo_operacion") or elemento.get("codigo") or "").strip().upper()
        elemento["usuarios"]=usuarios_por_trafo.get(codigo,[])
        elemento["medidas_auxiliares_excluidas"]=excluidos_por_trafo.get(codigo,0)
        elemento.update(estado_por_trafo.get(codigo,{}))
    data["version_sac_activos"] = 1
    return data

def _consultar_trazado(sub, fid, force_refresh=False, incluir_sac=True):
    data=GTechService().trazar_circuito(sub,fid,force_refresh=force_refresh)
    data["fecha_consulta"] = datetime.now().astimezone().isoformat(timespec="seconds")
    data["version_direcciones"] = 2
    data["version_sac_activos"] = 0
    return _completar_sac(data) if incluir_sac else data

def _guardar_cache(archivo, data):
    with _CIRCUITOS_CACHE_LOCK:
        archivo.parent.mkdir(parents=True, exist_ok=True)
        temporal = archivo.with_suffix(".tmp")
        datos_cache = {**data, "elementos": [{k: v for k, v in elemento.items() if k != "estado_simulado"} for elemento in data.get("elementos", [])]}
        with temporal.open("w", encoding="utf-8") as salida:
            json.dump(datos_cache, salida, ensure_ascii=False, separators=(",", ":"))
        os.replace(temporal, archivo)

def _aplicar_ampacidades(data):
    lineas=[e for e in data.get("elementos",[]) if e.get("g3e_fno")==19000]
    codigos={str(e.get("codigo_conductor") or "").strip().upper() for e in lineas}
    catalogo={x.codigo.strip().upper():x for x in CatalogoConductorCens.objects.filter(codigo__in=codigos)}
    subestaciones={str(e.get("subestacion") or data.get("subestacion") or "").strip().upper() for e in lineas}
    circuitos={str(e.get("circuito") or "").strip().upper() for e in lineas}
    configuraciones={
        (x.subestacion,x.circuito,x.codigo_conductor,x.g3e_fid):x.tipo_tendido
        for x in ConfiguracionTendidoCircuito.objects.filter(subestacion__in=subestaciones,circuito__in=circuitos,codigo_conductor__in=codigos)
    }
    for elemento in data.get("elementos",[]):
        if not elemento.get("subestacion") and data.get("subestacion"):
            elemento["subestacion"]=data["subestacion"]
        codigo=str(elemento.get("codigo_conductor") or "").strip().upper()
        item=catalogo.get(codigo)
        if item:
            for campo in ("ampacidad_a","ampacidad_ducto_a","ampacidad_aire_a","resistividad_ohm_mm2_m","resistencia_ohm_km","reactancia_ohm_km","gmr_mm","diametro_mm","temperatura_referencia_c","tension_nominal_kv","seccion_mm2"):
                valor=getattr(item,campo)
                if valor is not None:elemento[campo]=float(valor)
            elemento["ampacidad_catalogo_a"]=float(item.ampacidad_a) if item.ampacidad_a is not None else None
            clave=(str(elemento.get("subestacion") or data.get("subestacion") or "").strip().upper(),str(elemento.get("circuito") or "").strip().upper(),codigo,int(elemento.get("g3e_fid") or 0))
            tipo_tendido=configuraciones.get(clave)
            elemento["tipo_tendido"]=tipo_tendido or ""
            if tipo_tendido==ConfiguracionTendidoCircuito.TIPO_DUCTO and item.ampacidad_ducto_a is not None:
                elemento["ampacidad_a"]=float(item.ampacidad_ducto_a)
            elif tipo_tendido==ConfiguracionTendidoCircuito.TIPO_AIRE and item.ampacidad_aire_a is not None:
                elemento["ampacidad_a"]=float(item.ampacidad_aire_a)
            elemento["fabricante"]=item.fabricante
            elemento["familia_conductor"]=item.familia
            elemento["confianza_parametros"]=item.confianza
            elemento["observaciones_parametros"]=item.observaciones
            elemento["fuente_tecnica"]=item.fuente_tecnica
    return data


def inicio(request):
    try: subestaciones=GTechService().listar_subestaciones(); error=None
    except Exception as exc: subestaciones=[]; error=str(exc)
    try: subestaciones_medidas=sorted({m["subestacion"] for m in cargar_catalogo().get("medidas",[])})
    except Exception: subestaciones_medidas=[]
    return render(request,"desconexiones/inicio.html",{"subestaciones":subestaciones,"subestaciones_medidas":subestaciones_medidas,"error":error})


@require_POST
def api_cargar_kmz(request):
    archivo = request.FILES.get("archivo")
    if not archivo:
        return JsonResponse({"error": "Debe seleccionar un archivo KMZ o KML."}, status=400)
    if not archivo.name.lower().endswith((".kmz", ".kml")):
        return JsonResponse({"error": "El archivo debe tener extensión .kmz o .kml."}, status=400)
    try:
        nombre, geojson = leer_kmz(archivo)
        tipos = {}
        for feature in geojson["features"]:
            tipo = feature["geometry"]["type"]
            tipos[tipo] = tipos.get(tipo, 0) + 1
        return JsonResponse({"nombre": nombre, "geojson": geojson, "elementos": len(geojson["features"]), "tipos": tipos})
    except ValueError as exc:
        return JsonResponse({"error": str(exc)}, status=400)

@require_GET
def api_circuitos(request):
    sub=request.GET.get("subestacion","").strip().upper()
    if not sub:return JsonResponse({"error":"Debe seleccionar una subestación."},status=400)
    try:return JsonResponse({"subestacion":sub,"circuitos":GTechService().listar_circuitos(sub)})
    except Exception as exc:return JsonResponse({"error":str(exc)},status=500)

@require_GET
def api_trazado(request):
    sub=request.GET.get("subestacion","").strip().upper()
    try: fid=int(request.GET.get("fid",""))
    except ValueError:return JsonResponse({"error":"El interruptor seleccionado no es válido."},status=400)
    if not sub:return JsonResponse({"error":"Debe seleccionar una subestación."},status=400)
    try:
        archivo=_archivo_circuito(sub,fid)
        actualizar=request.GET.get("actualizar","").lower() in {"1","true","si","sí"}
        rapido=request.GET.get("rapido","").lower() in {"1","true","si","sí"}
        completar_sac=request.GET.get("completar_sac","").lower() in {"1","true","si","sí"}
        descarga_automatica=False
        if not archivo.exists() and not actualizar:
            # Un JSON ausente se recupera automáticamente: primero se guarda
            # la topología y SAC se completa después sin bloquear el mapa.
            actualizar=True
            rapido=not completar_sac
            completar_sac=False
            descarga_automatica=True
        if completar_sac:
            with _CIRCUITOS_CACHE_LOCK:
                with archivo.open("r",encoding="utf-8") as entrada:data=json.load(entrada)
            if data.get("version_sac_activos") != 1:
                _completar_sac(data)
                _guardar_cache(archivo,data)
            data["origen_datos"]="cache_sac"
        elif archivo.exists() and not actualizar:
            with _CIRCUITOS_CACHE_LOCK:
                with archivo.open("r",encoding="utf-8") as entrada:data=json.load(entrada)
            requiere_calibres=any(e.get("g3e_fno")==19000 and "calibre" not in e for e in data.get("elementos",[]))
            requiere_direcciones=data.get("version_direcciones") != 2
            requiere_sac_activos=data.get("version_sac_activos") != 1
            requiere_postes=data.get("version_postes") != 1
            requiere_topologia=data.get("version_topologia") != 3
            if requiere_calibres or requiere_direcciones or requiere_postes or requiere_topologia:
                data=_consultar_trazado(sub,fid,force_refresh=True)
                _guardar_cache(archivo,data)
                data["origen_datos"]="consulta"
            else:data["origen_datos"]="cache"
        else:
            data=_consultar_trazado(sub,fid,force_refresh=actualizar,incluir_sac=not rapido)
            _guardar_cache(archivo,data)
            data["origen_datos"]="consulta"
        _aplicar_ampacidades(data)
        data["descarga_automatica"]=descarga_automatica
        operaciones=request.session.get("operaciones_345",{})
        fids=[int(e["g3e_fid"]) for e in data["elementos"] if e.get("g3e_fid") is not None]
        vigentes={x.g3e_fid:x for x in EstadoOperativoVigente.objects.filter(g3e_fid__in=fids,habilitado=True,fecha_inicio__lte=timezone.localdate())}
        for e in data["elementos"]:
            # El JSON conserva la topologia, no una fotografia operativa antigua.
            # La posicion normal es el estado estable; las maniobras de la sesion
            # se aplican aparte mediante estado_simulado.
            condicion=vigentes.get(int(e["g3e_fid"]))
            base=condicion.estado if condicion else (e.get("estado_estable") or e.get("estado_operativo") or "CLOSED")
            if condicion:
                e["estado_operativo_vigente"]=condicion.estado
                e["condicion_operativa_vigente"]={"fecha_inicio":condicion.fecha_inicio.isoformat(),"observacion":condicion.observacion,"codigo":condicion.codigo}
            e["estado_simulado"]=operaciones.get(str(e["g3e_fid"]),base)
        return JsonResponse(data)
    except Exception as exc:return JsonResponse({"error":str(exc)},status=500)

@require_POST
def api_operar(request):
    try:
        data=json.loads(request.body);fid=str(int(data["g3e_fid"]));estado=str(data["estado"]).upper()
        if estado not in {"OPEN","CLOSED"}:raise ValueError("Estado inválido.")
        operaciones=request.session.get("operaciones_345",{});operaciones[fid]=estado
        request.session["operaciones_345"]=operaciones;request.session.modified=True
        return JsonResponse({"ok":True,"g3e_fid":int(fid),"estado":estado})
    except Exception as exc:return JsonResponse({"error":str(exc)},status=400)


@require_POST
def api_restaurar_operaciones(request):
    request.session["operaciones_345"]={}
    request.session.modified=True
    return JsonResponse({"ok":True,"restaurados":True})

@require_GET
def api_admin_estados_operativos(request):
    condiciones=[]
    for x in EstadoOperativoVigente.objects.all():
        condiciones.append({"id":x.id,"g3e_fid":x.g3e_fid,"codigo":x.codigo,"subestacion":x.subestacion,"circuito":x.circuito,"tipo_equipo":x.tipo_equipo,"fecha_inicio":x.fecha_inicio.isoformat(),"estado":x.estado,"habilitado":x.habilitado,"observacion":x.observacion})
    return JsonResponse({"condiciones":condiciones,"fecha_actual":timezone.localdate().isoformat()})

@require_POST
def api_admin_guardar_estado_operativo(request):
    try:
        data=json.loads(request.body)
        fid=int(data["g3e_fid"])
        if data.get("eliminar"):
            eliminados,_=EstadoOperativoVigente.objects.filter(g3e_fid=fid).delete()
            return JsonResponse({"ok":True,"eliminados":eliminados})
        estado=str(data.get("estado") or "").upper()
        if estado not in {"OPEN","CLOSED"}:raise ValueError("Seleccione un estado abierto o cerrado.")
        fecha=datetime.strptime(str(data.get("fecha_inicio") or ""),"%Y-%m-%d").date()
        defaults={
            "codigo":str(data.get("codigo") or ""),"subestacion":str(data.get("subestacion") or ""),
            "circuito":str(data.get("circuito") or ""),"tipo_equipo":str(data.get("tipo_equipo") or ""),
            "fecha_inicio":fecha,"estado":estado,"habilitado":bool(data.get("habilitado",True)),
            "observacion":str(data.get("observacion") or "").strip(),
            "actualizado_por":request.user if request.user.is_authenticated else None,
        }
        obj,_=EstadoOperativoVigente.objects.update_or_create(g3e_fid=fid,defaults=defaults)
        return JsonResponse({"ok":True,"id":obj.id,"g3e_fid":fid,"estado":estado,"habilitado":obj.habilitado})
    except Exception as exc:return JsonResponse({"error":str(exc)},status=400)


@require_POST
def api_exportar_maniobras_xlsx(request):
    """Genera un libro XLSX real; no HTML con una extensión de Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        data=json.loads(request.body);secciones=data.get("secciones") or []
        if not secciones:raise ValueError("No hay maniobras para exportar.")
        libro=Workbook();libro.remove(libro.active);usados=set()
        encabezados=["#","Fecha","Subestación","Celda","Dispositivo","Estado","Dirección aproximada","Maniobra"]
        for indice,seccion in enumerate(secciones,1):
            base="".join(c for c in str(seccion.get("titulo") or f"Protocolo {indice}") if c not in "[]:*?/\\")[:31] or f"Protocolo {indice}";titulo=base;n=2
            while titulo in usados:
                sufijo=f" {n}";titulo=base[:31-len(sufijo)]+sufijo;n+=1
            usados.add(titulo);hoja=libro.create_sheet(titulo);hoja.append(encabezados)
            for celda in hoja[1]:celda.font=Font(bold=True,color="FFFFFF");celda.fill=PatternFill("solid",fgColor="173A58")
            for fila in seccion.get("filas") or []:hoja.append(list(fila)[:len(encabezados)])
            hoja.freeze_panes="A2";hoja.auto_filter.ref=hoja.dimensions
            for columna,ancho in zip(hoja.columns,[7,14,24,20,28,16,42,90]):
                hoja.column_dimensions[columna[0].column_letter].width=ancho
                for celda in columna:celda.alignment=Alignment(vertical="top",wrap_text=True)
        salida=BytesIO();libro.save(salida)
        nombre="".join(c for c in str(data.get("nombre") or "maniobras-cdym") if c.isalnum() or c in "-_") or "maniobras-cdym"
        respuesta=HttpResponse(salida.getvalue(),content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        respuesta["Content-Disposition"]=f'attachment; filename="{nombre}.xlsx"'
        return respuesta
    except Exception as exc:return JsonResponse({"error":str(exc)},status=400)


def _datos_medida(objeto):
    if not objeto or not objeto.medida_dispositivo:
        return None
    return {
        "medida_subestacion": objeto.medida_subestacion,
        "medida_dispositivo": objeto.medida_dispositivo,
        "medida_fuente": objeto.medida_fuente,
        "nivel_kv": objeto.nivel_kv,
        "tipo_objeto": objeto.tipo_objeto,
    }


def _capacidad_barra(asignacion):
    if not asignacion or asignacion.capacidad_transformador_mva is None:
        return None
    return {
        "codigo": asignacion.gtech_codigo or asignacion.gtech_circuito or asignacion.gtech_fid,
        "capacidad": float(asignacion.capacidad_transformador_mva),
        "unidad_capacidad": "MVA",
    }


def _medida_barra_alimentador(subestacion, alimentador_fid):
    barras = GTechService().listar_barras_dispositivos(subestacion)
    barra = next((b for b in barras if any(int(d["g3e_fid"]) == alimentador_fid for d in b["dispositivos"])), None)
    if not barra:
        return None, "No se encontró la barra física conectada al alimentador.", None, None

    configuracion_barra = AsignacionMedidaEnergia.objects.filter(
        tipo_objeto=AsignacionMedidaEnergia.TIPO_BARRA,
        gtech_fid=barra["g3e_fid"],
    ).first()
    capacidad_configurada = _capacidad_barra(configuracion_barra)
    directa = configuracion_barra if configuracion_barra and configuracion_barra.medida_dispositivo else None
    if directa:
        lado = "baja" if float(barra["nivel_kv"]) == 13.8 else "alta"
        return _datos_medida(directa), f"Medida disponible por el lado de {lado} ({directa.nivel_kv:g} kV).", barra["g3e_fid"], capacidad_configurada

    relacion = RelacionBarraTransformacion.objects.filter(barra_secundaria_fid=barra["g3e_fid"]).first()
    if relacion and relacion.barra_primaria_fid:
        primaria = AsignacionMedidaEnergia.objects.filter(
            tipo_objeto=AsignacionMedidaEnergia.TIPO_BARRA,
            gtech_fid=relacion.barra_primaria_fid,
        ).exclude(medida_dispositivo="").first()
        if primaria:
            return _datos_medida(primaria), f"Medida disponible por el lado de alta ({primaria.nivel_kv:g} kV).", barra["g3e_fid"], capacidad_configurada or _capacidad_barra(primaria)

    ramal = RamalTransformadorManual.objects.filter(
        barra_secundaria_fid=barra["g3e_fid"]
    ).exclude(medida_dispositivo="").first()
    if ramal:
        return {
            "medida_subestacion": ramal.medida_subestacion,
            "medida_dispositivo": ramal.medida_dispositivo,
            "medida_fuente": ramal.medida_fuente,
            "nivel_kv": 34.5,
            "tipo_objeto": AsignacionMedidaEnergia.TIPO_BARRA,
        }, "Medida disponible por el lado de alta del transformador.", barra["g3e_fid"], capacidad_configurada

    secundaria = RelacionBarraTransformacion.objects.filter(barra_primaria_fid=barra["g3e_fid"]).first()
    if secundaria:
        medida_secundaria = AsignacionMedidaEnergia.objects.filter(
            tipo_objeto=AsignacionMedidaEnergia.TIPO_BARRA,
            gtech_fid=secundaria.barra_secundaria_fid,
        ).exclude(medida_dispositivo="").first()
        if medida_secundaria:
            return _datos_medida(medida_secundaria), f"Medida disponible por el lado de baja ({medida_secundaria.nivel_kv:g} kV).", barra["g3e_fid"], capacidad_configurada or _capacidad_barra(medida_secundaria)
    return None, "La barra no tiene una medida de potencia asignada por alta ni por baja.", barra["g3e_fid"], None


def _medidas_transformador_alimentador(subestacion, alimentador_fid):
    medida, origen, barra_fid, transformador = _medida_barra_alimentador(subestacion, alimentador_fid)
    if not barra_fid:
        return ([medida] if medida else []), origen, barra_fid, transformador, None
    membresia = BarraGrupoTransformador.objects.select_related("grupo").filter(barra_fid=barra_fid).first()
    if not membresia:
        return ([medida] if medida else []), origen, barra_fid, transformador, f"BARRA:{barra_fid}"
    grupo = membresia.grupo
    barras_gtech = GTechService().listar_barras_dispositivos(subestacion)
    por_fid = {int(b["g3e_fid"]): b for b in barras_gtech}
    medidas, origenes, vistas = [], [], set()
    for miembro in grupo.barras.all():
        fid_barra = int(miembro.barra_fid)
        if fid_barra == int(barra_fid):
            candidata, descripcion = medida, origen
        else:
            barra = por_fid.get(fid_barra)
            alimentador = next((d for d in (barra or {}).get("dispositivos", []) if int(d.get("g3e_fno", 0)) == 18800), None)
            if not alimentador:
                continue
            candidata, descripcion, _, _ = _medida_barra_alimentador(subestacion, int(alimentador["g3e_fid"]))
        if candidata:
            clave = (candidata.get("medida_subestacion"), candidata.get("medida_dispositivo"), candidata.get("medida_fuente"))
            if clave not in vistas:
                vistas.add(clave); medidas.append(candidata); origenes.append(f"Barra {fid_barra}: {descripcion}")
    capacidad = float(grupo.capacidad_mva) if grupo.capacidad_mva is not None else None
    transformador_grupo = {"codigo": grupo.nombre, "capacidad": capacidad, "unidad_capacidad": "MVA", "barras_fid": [b.barra_fid for b in grupo.barras.all()]}
    return medidas, " · ".join(origenes), barra_fid, transformador_grupo, f"TRAFO:{grupo.id}"

def _corregir_escala_potencias(potencias, corrientes, voltajes=None):
    """Normaliza P/Q/S y evalúa su coherencia con kV, A y factor de potencia."""
    if not potencias or not corrientes:
        return potencias

    def valores_por_instante(bloque, claves=None):
        salida = {}
        for serie in (bloque or {}).get("series", []):
            if claves and serie.get("clave") not in claves:
                continue
            for instante, valor in serie.get("puntos", []):
                if valor is not None:
                    salida.setdefault(str(instante), []).append(abs(float(valor)))
        return salida

    series = {s.get("clave"): s for s in potencias.get("series", [])}
    serie_s = series.get("S")
    if not serie_s:
        return potencias
    corrientes_t = valores_por_instante(corrientes, {"IR", "IS", "IT"})
    voltajes_t = valores_por_instante(voltajes, {"URS", "UST", "UTR"})
    try:
        nivel_kv = float(potencias.get("nivel_kv") or corrientes.get("nivel_kv") or 0)
    except (TypeError, ValueError):
        nivel_kv = 0

    comparaciones = []
    for instante, valor_s in serie_s.get("puntos", []):
        fases_i = corrientes_t.get(str(instante), [])
        if not fases_i or not valor_s:
            continue
        fases_v = voltajes_t.get(str(instante), [])
        tension_kv = sum(fases_v) / len(fases_v) if fases_v else nivel_kv
        corriente_a = sum(fases_i) / len(fases_i)
        aparente_electrica = (3 ** 0.5) * tension_kv * corriente_a / 1000.0
        if aparente_electrica > 0:
            comparaciones.append(aparente_electrica / abs(float(valor_s)))
    if not comparaciones:
        return potencias
    ordenadas = sorted(comparaciones)
    relacion = ordenadas[len(ordenadas) // 2]
    factor = 1000.0 if 100 <= relacion <= 10000 else (0.001 if 0.0001 <= relacion <= 0.01 else 1.0)
    if factor != 1.0:
        for serie in potencias.get("series", []):
            if serie.get("clave") in {"P", "Q", "S"}:
                serie["puntos"] = [[p[0], float(p[1]) * factor] for p in serie.get("puntos", [])]
        potencias["escala_corregida_por_coherencia"] = True
    potencias["factor_correccion"] = factor

    p_t = {str(t): abs(float(v)) for t, v in series.get("P", {}).get("puntos", []) if v is not None}
    q_t = {str(t): abs(float(v)) for t, v in series.get("Q", {}).get("puntos", []) if v is not None}
    factores_potencia = []
    correlaciones = []
    puntos_s = []
    for instante, valor_s in serie_s.get("puntos", []):
        clave = str(instante)
        p = p_t.get(clave)
        q = q_t.get(clave)
        s_pq = (p * p + q * q) ** 0.5 if p is not None and q is not None else abs(float(valor_s))
        puntos_s.append([instante, s_pq])
        if s_pq > 0 and p is not None:
            factores_potencia.append(min(1.0, p / s_pq))
        fases_i = corrientes_t.get(clave, [])
        if fases_i:
            fases_v = voltajes_t.get(clave, [])
            tension_kv = sum(fases_v) / len(fases_v) if fases_v else nivel_kv
            s_electrica = (3 ** 0.5) * tension_kv * (sum(fases_i) / len(fases_i)) / 1000.0
            if s_electrica > 0 and s_pq > 0:
                correlaciones.append(min(s_electrica, s_pq) / max(s_electrica, s_pq))
    serie_s["puntos"] = puntos_s
    if factores_potencia:
        fp = sorted(factores_potencia)
        potencias["factor_potencia_mediano"] = fp[len(fp) // 2]
        potencias["factor_potencia_minimo"] = min(fp)
        potencias["factor_potencia_maximo"] = max(fp)
    if correlaciones:
        co = sorted(correlaciones)
        mediana = co[len(co) // 2]
        potencias["correlacion_kv_a_potencia"] = mediana
        potencias["calidad_correlacion"] = "BUENA" if mediana >= 0.85 else ("REVISAR" if mediana >= 0.65 else "DEFICIENTE")
    return potencias

def _potencia_transformador_agrupada(servicio, medidas, fecha, nombre):
    bloques = []
    for medida in medidas:
        potencia = servicio.serie_dia(medida, fecha, VARIABLES["POTENCIAS"])
        corrientes = servicio.serie_dia(medida, fecha, VARIABLES["CORRIENTE"])
        voltajes = servicio.serie_dia(medida, fecha, VARIABLES["VOLTAJE"])
        _corregir_escala_potencias(potencia, corrientes, voltajes)
        bloques.append(potencia)
    if not bloques:
        return None
    if len(bloques) == 1:
        return bloques[0]
    acumulado = {"P": {}, "Q": {}}
    for bloque in bloques:
        for serie in bloque.get("series", []):
            clave = serie.get("clave")
            if clave not in acumulado:
                continue
            for instante, valor in serie.get("puntos", []):
                acumulado[clave][str(instante)] = acumulado[clave].get(str(instante), 0.0) + abs(float(valor))
    instantes = sorted(set(acumulado["P"]) | set(acumulado["Q"]))
    puntos_p = [[t, acumulado["P"].get(t, 0.0)] for t in instantes]
    puntos_q = [[t, acumulado["Q"].get(t, 0.0)] for t in instantes]
    puntos_s = [[t, (acumulado["P"].get(t, 0.0) ** 2 + acumulado["Q"].get(t, 0.0) ** 2) ** 0.5] for t in instantes]
    return {**bloques[0], "dispositivo": nombre, "cantidad_registros": len(instantes), "medidas_agrupadas": len(bloques), "series": [
        {"clave": "P", "nombre": "P activa", "unidad": "MW", "puntos": puntos_p},
        {"clave": "Q", "nombre": "Q reactiva", "unidad": "MVAr", "puntos": puntos_q},
        {"clave": "S", "nombre": "S aparente", "unidad": "MVA", "puntos": puntos_s},
    ]}

_PATIOS_ENTRADAS={46410804:"BELEN",46410772:"SANMATEO"}
_PATIOS_SALIDA_FID=46407122

def _estado_actual_patios(request, fid):
    operaciones=request.session.get("operaciones_345",{})
    if str(fid) in operaciones:return operaciones[str(fid)]
    condicion=EstadoOperativoVigente.objects.filter(g3e_fid=fid,habilitado=True,fecha_inicio__lte=timezone.localdate()).first()
    return condicion.estado if condicion else "CLOSED"

def _sumar_bloques_medida(base, adicionales, claves_sumables):
    if not base:return base
    salida=json.loads(json.dumps(base))
    for serie in salida.get("series",[]):
        if serie.get("clave") not in claves_sumables:continue
        extras=[next((s for s in bloque.get("series",[]) if s.get("clave")==serie.get("clave")),None) for bloque in adicionales if bloque]
        extras=[s for s in extras if s]
        serie["puntos"]=[[p[0],abs(float(p[1] or 0))+sum(abs(float(s.get("puntos",[])[i][1] or 0)) for s in extras if i<len(s.get("puntos",[])))] for i,p in enumerate(serie.get("puntos",[]))]
    return salida

def _serie_componente_patios(servicio, medida, fecha_base, variable, modo, request):
    fecha=fecha_base;seleccion=None
    if modo=="HISTORICO":
        seleccion=servicio.seleccionar_dia_maxima_corriente(
            medida,
            int(request.GET.get("dia_semana","0")),
            request.GET.get("tipo_dia","ORDINARIO"),
            float(request.GET.get("delta","20")),
        )
        fecha=servicio.validar_fecha(seleccion["fecha"])
    bloque=servicio.serie_dia(medida,fecha,VARIABLES[variable])
    return bloque,{"dispositivo":medida.get("medida_dispositivo"),"fecha":fecha.isoformat(),"seleccion_historica":seleccion}

def _aplicar_nodo_patios(request, servicio, alimentador_fid, fecha, modo, variable, alimentador, potencias, corrientes, voltajes):
    if alimentador_fid not in _PATIOS_ENTRADAS:return alimentador,potencias,corrientes,voltajes,None
    estados={fid:_estado_actual_patios(request,fid) for fid in _PATIOS_ENTRADAS}
    activas=[fid for fid,estado in estados.items() if estado=="CLOSED"]
    detalle={"tipo":"NODO_345_PATIOS","entradas":estados,"entrada_activa":activas[0] if len(activas)==1 else None,"origen":_PATIOS_ENTRADAS.get(activas[0]) if len(activas)==1 else None,"aplicado":False,"componentes":[]}
    if len(activas)!=1:
        detalle["advertencia"]="Las entradas de BELEN y SANMATEO estan simultaneamente cerradas." if len(activas)>1 else "La barra de LOS PATIOS no tiene una entrada cerrada."
        return alimentador,potencias,corrientes,voltajes,detalle
    if activas[0]!=alimentador_fid:return alimentador,potencias,corrientes,voltajes,detalle
    salida=AsignacionMedidaEnergia.objects.filter(tipo_objeto=AsignacionMedidaEnergia.TIPO_ALIMENTADOR,gtech_fid=_PATIOS_SALIDA_FID).exclude(medida_dispositivo="").first()
    ramal=RamalTransformadorManual.objects.filter(subestacion="PATIOS",barra_primaria_fid=46407097).exclude(medida_dispositivo="").first()
    medidas=[]
    if salida:medidas.append(_datos_medida(salida))
    if ramal:medidas.append({"medida_subestacion":ramal.medida_subestacion,"medida_dispositivo":ramal.medida_dispositivo,"medida_fuente":ramal.medida_fuente,"nivel_kv":34.5,"tipo_objeto":AsignacionMedidaEnergia.TIPO_BARRA})
    adicionales_variable=[];adicionales_p=[];adicionales_i=[]
    for medida in medidas:
        bloque_variable,info=_serie_componente_patios(servicio,medida,fecha,variable,modo,request)
        bloque_p,_=_serie_componente_patios(servicio,medida,fecha,"POTENCIAS",modo,request)
        bloque_i,_=_serie_componente_patios(servicio,medida,fecha,"CORRIENTE",modo,request)
        bloque_v,_=_serie_componente_patios(servicio,medida,fecha,"VOLTAJE",modo,request)
        _corregir_escala_potencias(bloque_p,bloque_i,bloque_v)
        adicionales_variable.append(bloque_variable);adicionales_p.append(bloque_p);adicionales_i.append(bloque_i);detalle["componentes"].append(info)
    if variable=="CORRIENTE":alimentador=_sumar_bloques_medida(alimentador,adicionales_variable,{"IR","IS","IT"})
    elif variable=="POTENCIAS":alimentador=_sumar_bloques_medida(alimentador,adicionales_variable,{"P","Q","S"})
    potencias=_sumar_bloques_medida(potencias,adicionales_p,{"P","Q","S"})
    corrientes=_sumar_bloques_medida(corrientes,adicionales_i,{"IR","IS","IT"})
    detalle["aplicado"]=True
    detalle["formula"]="Entrada activa + P_IL30 + PATIT40"
    return alimentador,potencias,corrientes,voltajes,detalle

@require_GET
def api_grafica_medidas(request):
    subestacion = request.GET.get("subestacion", "").strip().upper()
    variable = request.GET.get("variable", "VOLTAJE").strip().upper()
    try:
        alimentador_fid = int(request.GET.get("fid", ""))
        if not subestacion:
            raise ValueError("Debe indicar la subestación.")
        if variable not in VARIABLES:
            raise ValueError("La variable seleccionada no es válida.")

        asignacion = AsignacionMedidaEnergia.objects.filter(
            tipo_objeto=AsignacionMedidaEnergia.TIPO_ALIMENTADOR,
            gtech_fid=alimentador_fid,
        ).exclude(medida_dispositivo="").first()
        if not asignacion:
            return JsonResponse({"error": "El alimentador no tiene una medida histórica asignada."}, status=404)

        servicio = SeriesMedidasService()
        fecha_minima, fecha_maxima = servicio.rango_global()
        modo = request.GET.get("modo", "FECHA").strip().upper()
        seleccion_historica = None
        if modo == "HISTORICO":
            dia_semana = int(request.GET.get("dia_semana", "0"))
            tipo_dia = request.GET.get("tipo_dia", "ORDINARIO")
            delta = float(request.GET.get("delta", "20"))
            seleccion_historica = servicio.seleccionar_dia_maxima_corriente(_datos_medida(asignacion), dia_semana, tipo_dia, delta)
            fecha = servicio.validar_fecha(seleccion_historica["fecha"])
        else:
            fecha = servicio.validar_fecha(request.GET.get("fecha")) or servicio.validar_fecha(fecha_maxima)
        alimentador = servicio.serie_dia(_datos_medida(asignacion), fecha, VARIABLES[variable])
        alimentador_potencias = alimentador if variable == "POTENCIAS" else servicio.serie_dia(_datos_medida(asignacion), fecha, VARIABLES["POTENCIAS"])
        alimentador_corrientes = alimentador if variable == "CORRIENTE" else servicio.serie_dia(_datos_medida(asignacion), fecha, VARIABLES["CORRIENTE"])
        alimentador_voltajes = alimentador if variable == "VOLTAJE" else servicio.serie_dia(_datos_medida(asignacion), fecha, VARIABLES["VOLTAJE"])
        _corregir_escala_potencias(alimentador_potencias, alimentador_corrientes, alimentador_voltajes)
        alimentador,alimentador_potencias,alimentador_corrientes,alimentador_voltajes,nodo_345=_aplicar_nodo_patios(request,servicio,alimentador_fid,fecha,modo,variable,alimentador,alimentador_potencias,alimentador_corrientes,alimentador_voltajes)
        solo_alimentador = request.GET.get("solo_alimentador", "").lower() in {"1", "true", "si", "sí"}
        if solo_alimentador:
            medidas_barra, origen_barra, barra_fid, transformador, transformador_key = [], "Consulta de bloque de carga de 34,5 kV.", None, None, None
        else:
            medidas_barra, origen_barra, barra_fid, transformador, transformador_key = _medidas_transformador_alimentador(subestacion, alimentador_fid)
        barra = _potencia_transformador_agrupada(
            servicio,
            medidas_barra,
            fecha,
            transformador.get("codigo") if transformador else "",
        )
        return JsonResponse({
            "fecha": fecha.isoformat(),
            "fecha_minima": fecha_minima,
            "fecha_maxima": fecha_maxima,
            "variable": variable,
            "modo_seleccion": modo,
            "seleccion_historica": seleccion_historica,
            "set_proteccion_a": float(asignacion.set_proteccion_a) if asignacion.set_proteccion_a is not None else None,
            "alimentador": alimentador,
            "alimentador_potencias": alimentador_potencias,
            "alimentador_corrientes": alimentador_corrientes,
            "alimentador_voltajes": alimentador_voltajes,
            "barra": barra,
            "barra_fid": barra_fid,
            "transformador_key": transformador_key,
            "transformador": transformador,
            "origen_barra": origen_barra,
            "nodo_345": nodo_345,
        })
    except (ValueError, FileNotFoundError) as exc:
        return JsonResponse({"error": str(exc)}, status=400)
    except Exception as exc:
        return JsonResponse({"error": str(exc)}, status=500)


@require_GET
def api_admin_barras(request):
    sub=request.GET.get("subestacion","").strip().upper()
    if not sub:return JsonResponse({"error":"Debe seleccionar una subestación."},status=400)
    try:
        barras=GTechService().listar_barras_dispositivos(sub)
        sub_medida=request.GET.get("medida_subestacion",sub).strip().upper()
        medidas=medidas_subestacion(sub_medida)
        fids=[b["g3e_fid"] for b in barras]+[d["g3e_fid"] for b in barras for d in b["dispositivos"]]
        guardadas={(a.tipo_objeto,a.gtech_fid):a for a in AsignacionMedidaEnergia.objects.filter(gtech_fid__in=fids)}
        relaciones={r.barra_secundaria_fid:{"barra_primaria_fid":r.barra_primaria_fid,"origen_tipo":r.origen_tipo,"nivel_primario_kv":r.nivel_primario_kv} for r in RelacionBarraTransformacion.objects.filter(barra_secundaria_fid__in=[b["g3e_fid"] for b in barras])}
        membresias={m.barra_fid:m for m in BarraGrupoTransformador.objects.select_related("grupo").filter(barra_fid__in=[b["g3e_fid"] for b in barras])}
        def preparar(obj,tipo,nivel):
            guardada=guardadas.get((tipo,obj["g3e_fid"]));exacta=coincidencia_preferida(obj,medidas,nivel)
            seleccion=None
            if guardada and guardada.medida_dispositivo:seleccion={"subestacion":guardada.medida_subestacion,"interruptor":guardada.medida_dispositivo,"fuente":guardada.medida_fuente,"nivel_kv":guardada.nivel_kv}
            elif exacta:seleccion=exacta
            obj["tipo_objeto"]=tipo;obj["coincidencia_exacta"]=bool(exacta);obj["preseleccion_automatica"]=not bool(guardada) and bool(exacta);obj["seleccion"]=seleccion
            sugerido,origen_set=set_sugerido(obj) if not guardada and tipo==AsignacionMedidaEnergia.TIPO_ALIMENTADOR and float(nivel)==13.8 else (None,"")
            obj["set_proteccion_a"]=float(guardada.set_proteccion_a) if guardada and guardada.set_proteccion_a is not None else sugerido
            obj["set_origen"]="GUARDADO" if guardada else origen_set
            obj["capacidad_transformador_mva"]=float(guardada.capacidad_transformador_mva) if guardada and guardada.capacidad_transformador_mva is not None else None
            obj["funcion_electrica"]=guardada.funcion_electrica if guardada and guardada.funcion_electrica else ("ALIMENTADOR" if float(nivel)==13.8 else "LINEA_345")
        for barra in barras:
            preparar(barra,AsignacionMedidaEnergia.TIPO_BARRA,barra["nivel_kv"])
            if float(barra["nivel_kv"])==13.8:
                barra["origen_transformador"]=relaciones.get(barra["g3e_fid"])
                membresia=membresias.get(barra["g3e_fid"])
                barra["grupo_transformador"]={"id":membresia.grupo_id,"nombre":membresia.grupo.nombre,"capacidad_mva":float(membresia.grupo.capacidad_mva) if membresia.grupo.capacidad_mva is not None else None} if membresia else {"id":None,"nombre":f"{sub}_BARRA_{barra['g3e_fid']}","capacidad_mva":barra.get("capacidad_transformador_mva")}
            for dispositivo in barra["dispositivos"]:preparar(dispositivo,AsignacionMedidaEnergia.TIPO_ALIMENTADOR,barra["nivel_kv"])
        ramales=[{"id":r.id,"tipo_manual":r.tipo_manual,"nombre":r.nombre,"barra_primaria_fid":r.barra_primaria_fid,"barra_secundaria_fid":r.barra_secundaria_fid,"seleccion":{"subestacion":r.medida_subestacion,"interruptor":r.medida_dispositivo,"fuente":r.medida_fuente,"nivel_kv":34.5} if r.medida_dispositivo else None} for r in RamalTransformadorManual.objects.filter(subestacion=sub)]
        grupos=[{"id":g.id,"nombre":g.nombre,"capacidad_mva":float(g.capacidad_mva) if g.capacidad_mva is not None else None,"barras_fid":[m.barra_fid for m in g.barras.all()]} for g in GrupoTransformadorBarra.objects.filter(subestacion=sub).prefetch_related("barras")]
        return JsonResponse({"subestacion":sub,"medida_subestacion":sub_medida,"barras":barras,"medidas":medidas,"grupos_transformadores":grupos,"ramales_manuales":ramales})
    except Exception as exc:return JsonResponse({"error":str(exc)},status=500)

@require_POST
def api_admin_guardar_medidas(request):
    try:
        payload=json.loads(request.body);items=payload.get("asignaciones",[])
        with transaction.atomic():
            for item in items:
                tipo=str(item.get("tipo_objeto","")).upper();fid=int(item["gtech_fid"]);medida=item.get("medida")
                if tipo not in {AsignacionMedidaEnergia.TIPO_BARRA,AsignacionMedidaEnergia.TIPO_ALIMENTADOR}:raise ValueError("Tipo de objeto inválido.")
                if not medida:
                    funcion=str(item.get("funcion_electrica", ""));capacidad=item.get("capacidad_transformador_mva")
                    if (tipo==AsignacionMedidaEnergia.TIPO_ALIMENTADOR and funcion) or (tipo==AsignacionMedidaEnergia.TIPO_BARRA and capacidad):
                        AsignacionMedidaEnergia.objects.update_or_create(tipo_objeto=tipo,gtech_fid=fid,defaults={"gtech_codigo":str(item.get("gtech_codigo", "")),"gtech_circuito":str(item.get("gtech_circuito", "")),"subestacion":str(item.get("subestacion", "")).upper(),"nivel_kv":float(item["nivel_kv"]),"medida_subestacion":"","medida_dispositivo":"","medida_fuente":"","coincidencia_exacta":False,"funcion_electrica":funcion,"set_proteccion_a":item.get("set_proteccion_a") or None,"capacidad_transformador_mva":item.get("capacidad_transformador_mva") or None})
                    else:AsignacionMedidaEnergia.objects.filter(tipo_objeto=tipo,gtech_fid=fid).delete()
                    continue
                nombres={normalizar(item.get(k)) for k in ("gtech_codigo","gtech_circuito","gtech_marcacion")};nombres.discard("")
                exacta=normalizar(medida.get("interruptor")) in nombres
                AsignacionMedidaEnergia.objects.update_or_create(tipo_objeto=tipo,gtech_fid=fid,defaults={"gtech_codigo":str(item.get("gtech_codigo", "")),"gtech_circuito":str(item.get("gtech_circuito", "")),"subestacion":str(item.get("subestacion", "")).upper(),"nivel_kv":float(item["nivel_kv"]),"medida_subestacion":str(medida.get("subestacion", "")),"medida_dispositivo":str(medida["interruptor"]),"medida_fuente":str(medida.get("fuente", "")),"coincidencia_exacta":exacta,"funcion_electrica":str(item.get("funcion_electrica", "")),"set_proteccion_a":item.get("set_proteccion_a") or None,"capacidad_transformador_mva":item.get("capacidad_transformador_mva") or None})
            for relacion in payload.get("relaciones_transformacion",[]):
                secundaria=int(relacion["barra_secundaria_fid"]);origen=str(relacion.get("origen_tipo","")).upper();primaria=relacion.get("barra_primaria_fid");nivel=relacion.get("nivel_primario_kv")
                if origen:RelacionBarraTransformacion.objects.update_or_create(barra_secundaria_fid=secundaria,defaults={"barra_primaria_fid":int(primaria) if primaria else None,"origen_tipo":origen,"nivel_primario_kv":float(nivel) if nivel else None,"subestacion":str(relacion.get("subestacion","")).upper()})
                else:RelacionBarraTransformacion.objects.filter(barra_secundaria_fid=secundaria).delete()
            grupos_payload=payload.get("grupos_transformacion",[])
            for item in grupos_payload:
                barra_fid=int(item["barra_fid"]);nombre=str(item.get("nombre","")).strip().upper();capacidad=item.get("capacidad_mva")
                if not nombre:raise ValueError(f"La barra {barra_fid} debe tener un transformador asociado.")
                sub_grupo=str(item.get("subestacion","")).strip().upper()
                grupo,_=GrupoTransformadorBarra.objects.update_or_create(subestacion=sub_grupo,nombre=nombre,defaults={"capacidad_mva":float(capacidad) if capacidad not in (None,"") else None})
                BarraGrupoTransformador.objects.update_or_create(barra_fid=barra_fid,defaults={"grupo":grupo})
            sub_grupos=str(payload.get("subestacion","")).strip().upper()
            if sub_grupos:GrupoTransformadorBarra.objects.filter(subestacion=sub_grupos,barras__isnull=True).delete()
            ramales_payload=payload.get("ramales_manuales",[]);ids=[]
            for ramal in ramales_payload:
                nombre=str(ramal.get("nombre","")).strip().upper()
                if not nombre:continue
                medida=ramal.get("medida") or {};rid=ramal.get("id")
                defaults={"tipo_manual":str(ramal.get("tipo_manual","TRANSFORMADOR")).upper(),"subestacion":str(ramal.get("subestacion","")).upper(),"nombre":nombre,"barra_primaria_fid":int(ramal["barra_primaria_fid"]),"barra_secundaria_fid":int(ramal["barra_secundaria_fid"]) if ramal.get("barra_secundaria_fid") else None,"medida_subestacion":str(medida.get("subestacion","")),"medida_dispositivo":str(medida.get("interruptor","")),"medida_fuente":str(medida.get("fuente",""))}
                if rid:obj,_=RamalTransformadorManual.objects.update_or_create(id=int(rid),defaults=defaults)
                else:obj,_=RamalTransformadorManual.objects.update_or_create(subestacion=defaults["subestacion"],nombre=nombre,defaults=defaults)
                ids.append(obj.id)
            sub_payload=str(payload.get("subestacion","")).upper()
            if sub_payload:RamalTransformadorManual.objects.filter(subestacion=sub_payload).exclude(id__in=ids).delete()
        return JsonResponse({"ok":True,"guardadas":len(items),"relaciones":len(payload.get("relaciones_transformacion",[])),"grupos_transformacion":len(payload.get("grupos_transformacion",[])),"ramales":len(payload.get("ramales_manuales",[]))})
    except Exception as exc:return JsonResponse({"error":str(exc)},status=400)



@require_GET
def api_aprendizaje_protocolo(request):
    firma=str(request.GET.get("firma","")).strip().lower()
    perfil=PerfilAprendizajeManiobras.objects.filter(nombre="GLOBAL").first()
    preferencias=perfil.preferencias if perfil else {}
    if not firma:return JsonResponse({"aprendizaje":None,"preferencias":preferencias})
    item=AprendizajeProtocolo.objects.filter(firma=firma).first()
    if not item:return JsonResponse({"aprendizaje":None,"preferencias":preferencias})
    return JsonResponse({"aprendizaje":{"firma":item.firma,"tipo_cambio":item.tipo_cambio,"motivo":item.motivo,"protocolo":item.protocolo,"contexto":item.contexto,"actualizado_en":item.actualizado_en.isoformat()},"preferencias":preferencias})

@require_POST
def api_guardar_aprendizaje_protocolo(request):
    try:
        payload=json.loads(request.body);firma=str(payload.get("firma","")).strip().lower();motivo=str(payload.get("motivo","")).strip();protocolo=payload.get("protocolo",[])
        if len(firma)!=64 or any(c not in "0123456789abcdef" for c in firma):raise ValueError("La firma del escenario no es valida.")
        if not motivo:raise ValueError("Debe indicar el motivo tecnico del cambio.")
        if not isinstance(protocolo,list) or len(protocolo)>500:raise ValueError("El protocolo aprendido no es valido.")
        anterior=AprendizajeProtocolo.objects.filter(firma=firma).first();contexto=payload.get("contexto",{}) if isinstance(payload.get("contexto",{}),dict) else {};historial=list((anterior.contexto or {}).get("historial",[])) if anterior else [];historial.append({"tipo_cambio":str(payload.get("tipo_cambio","EDICION"))[:20],"motivo":motivo[:2000],"fecha":datetime.now().astimezone().isoformat(timespec="seconds")});contexto["historial"]=historial[-100:];usuario=request.user if request.user.is_authenticated else None;tipo=str(payload.get("tipo_cambio","EDICION"))[:20]
        with transaction.atomic():
            EventoAprendizajeProtocolo.objects.create(firma=firma,tipo_cambio=tipo,motivo=motivo[:2000],protocolo_anterior=anterior.protocolo if anterior else [],protocolo_corregido=protocolo,contexto=contexto,creado_por=usuario)
            item,_=AprendizajeProtocolo.objects.update_or_create(firma=firma,defaults={"tipo_cambio":tipo,"motivo":motivo[:2000],"protocolo":protocolo,"contexto":contexto,"actualizado_por":usuario})
            preferencias=reconstruir_preferencias(EventoAprendizajeProtocolo.objects.all().order_by("creado_en"))
            perfil,_=PerfilAprendizajeManiobras.objects.update_or_create(nombre="GLOBAL",defaults={"preferencias":preferencias,"ejemplos":preferencias["ejemplos"],"version":preferencias["version"]})
        return JsonResponse({"ok":True,"id":item.id,"evento_id":EventoAprendizajeProtocolo.objects.filter(firma=firma).latest("creado_en").id,"ejemplos":perfil.ejemplos,"actualizado_en":item.actualizado_en.isoformat()})
    except Exception as exc:return JsonResponse({"error":str(exc)},status=400)
@require_GET
def api_admin_paralelos(request):
    items=ParaleloCeldaPermitido.objects.all()
    return JsonResponse({"paralelos":[{"id":x.id,"subestacion_a":x.subestacion_a,"celda_a_fid":x.celda_a_fid,"celda_a_codigo":x.celda_a_codigo,"subestacion_b":x.subestacion_b,"celda_b_fid":x.celda_b_fid,"celda_b_codigo":x.celda_b_codigo,"nivel_kv":x.nivel_kv,"observacion":x.observacion,"activo":x.activo} for x in items]})

@require_POST
def api_admin_guardar_paralelos(request):
    try:
        payload=json.loads(request.body);items=payload.get("paralelos",[]);ids=[]
        with transaction.atomic():
            for item in items:
                fa=int(item["celda_a_fid"]);fb=int(item["celda_b_fid"])
                if fa==fb:raise ValueError("Las dos celdas del paralelo deben ser diferentes.")
                nivel=float(item.get("nivel_kv") or 13.8)
                if nivel not in {13.8,34.5}:raise ValueError("Solo se permiten paralelos de 13,8 o 34,5 kV.")
                defaults={"subestacion_a":str(item["subestacion_a"]),"celda_a_codigo":str(item["celda_a_codigo"]),"subestacion_b":str(item["subestacion_b"]),"celda_b_codigo":str(item["celda_b_codigo"]),"nivel_kv":nivel,"observacion":str(item.get("observacion","")).strip(),"activo":bool(item.get("activo",True))}
                a,b=sorted((fa,fb));obj,_=ParaleloCeldaPermitido.objects.update_or_create(celda_a_fid=a,celda_b_fid=b,defaults={**defaults,"celda_a_fid":fa,"celda_b_fid":fb})
                ids.append(obj.id)
            ParaleloCeldaPermitido.objects.exclude(id__in=ids).delete()
        return JsonResponse({"ok":True,"guardados":len(ids)})
    except Exception as exc:return JsonResponse({"error":str(exc)},status=400)

@require_GET
def api_admin_calibres(request):
    items=CatalogoConductorCens.objects.all()
    numericos=("tension_nominal_kv","seccion_mm2","ampacidad_a","ampacidad_ducto_a","ampacidad_aire_a","resistividad_ohm_mm2_m","resistencia_ohm_km","reactancia_ohm_km","gmr_mm","diametro_mm","temperatura_referencia_c")
    conductores=[]
    for x in items:
        item={campo:getattr(x,campo) for campo in ("codigo","descripcion","material","calibre","aislamiento","configuracion","familia","origen_ampacidad","fabricante","origen_parametros","confianza","observaciones","fuente_tecnica")}
        item.update({campo:float(getattr(x,campo)) if getattr(x,campo) is not None else None for campo in numericos})
        conductores.append(item)
    return JsonResponse({"conductores":conductores})

@require_POST
def api_admin_guardar_ampacidades(request):
    try:
        payload=json.loads(request.body);actualizados=0
        with transaction.atomic():
            for item in payload.get("conductores",[]):
                campos_numericos=("ampacidad_a","ampacidad_ducto_a","ampacidad_aire_a","resistividad_ohm_mm2_m","resistencia_ohm_km","reactancia_ohm_km","gmr_mm","diametro_mm","temperatura_referencia_c")
                defaults={campo:(float(item[campo]) if item.get(campo) not in (None,"") else None) for campo in campos_numericos if campo in item}
                defaults.update({campo:str(item.get(campo,"")).strip() for campo in ("fabricante","fuente_tecnica") if campo in item})
                actualizados+=CatalogoConductorCens.objects.filter(codigo=str(item.get("codigo","")).strip()).update(**defaults)
        return JsonResponse({"ok":True,"actualizados":actualizados})
    except Exception as exc:return JsonResponse({"error":str(exc)},status=400)

@require_POST
def api_guardar_tendido_conductor(request):
    try:
        payload=json.loads(request.body)
        subestacion=str(payload.get("subestacion") or "").strip().upper()
        circuito=str(payload.get("circuito") or "").strip().upper()
        codigo=str(payload.get("codigo_conductor") or "").strip().upper()
        tipo=str(payload.get("tipo_tendido") or "").strip().upper()
        fids=sorted({int(fid) for fid in payload.get("g3e_fids",[]) if str(fid).strip()})
        if not subestacion or not circuito or not codigo:
            raise ValueError("Faltan la subestacion, el circuito o el codigo del conductor.")
        if tipo not in {ConfiguracionTendidoCircuito.TIPO_DUCTO,ConfiguracionTendidoCircuito.TIPO_AIRE}:
            raise ValueError("Seleccione ducto/subterraneo o aire.")
        if not fids:raise ValueError("No se identificaron las lineas conectadas del tramo.")
        conductor=CatalogoConductorCens.objects.filter(codigo=codigo).first()
        if conductor is None:raise ValueError("El conductor no existe en el catalogo tecnico.")
        campo="ampacidad_ducto_a" if tipo==ConfiguracionTendidoCircuito.TIPO_DUCTO else "ampacidad_aire_a"
        ampacidad=getattr(conductor,campo)
        if ampacidad is None:raise ValueError("El catalogo no tiene ampacidad para el tipo de tendido seleccionado.")
        with transaction.atomic():
            for fid in fids:
                ConfiguracionTendidoCircuito.objects.update_or_create(
                    subestacion=subestacion,circuito=circuito,codigo_conductor=codigo,g3e_fid=fid,
                    defaults={"tipo_tendido":tipo},
                )
        return JsonResponse({"ok":True,"subestacion":subestacion,"circuito":circuito,"codigo_conductor":codigo,"g3e_fids":fids,"tipo_tendido":tipo,"ampacidad_a":float(ampacidad)})
    except Exception as exc:return JsonResponse({"error":str(exc)},status=400)

def _config_maniobras_json(config):
    return {
        "corriente_max_apertura_aisladero_a":float(config.corriente_max_apertura_aisladero_a),
        "corriente_max_cierre_aisladero_a":float(config.corriente_max_cierre_aisladero_a),
        "usuarios_totales_cens":config.usuarios_totales_cens,
        "maniobra_inicio_desenergizacion":config.maniobra_inicio_desenergizacion,
        "maniobra_fin_desenergizacion":config.maniobra_fin_desenergizacion,
        "maniobra_inicio_energizacion":config.maniobra_inicio_energizacion,
        "maniobra_fin_energizacion":config.maniobra_fin_energizacion,
    }

@require_GET
def api_admin_reglas_maniobra(request):
    config=ConfiguracionManiobras.actual()
    return JsonResponse(_config_maniobras_json(config))


@require_POST
def api_admin_guardar_reglas_maniobra(request):
    try:
        data=json.loads(request.body)
        apertura=float(data.get("corriente_max_apertura_aisladero_a"))
        cierre=float(data.get("corriente_max_cierre_aisladero_a"))
        if apertura<0 or cierre<0:
            raise ValueError("Las corrientes máximas no pueden ser negativas.")
        if apertura>10000 or cierre>10000:
            raise ValueError("Las corrientes máximas deben ser menores o iguales a 10000 A.")
        config=ConfiguracionManiobras.actual()
        config.corriente_max_apertura_aisladero_a=apertura
        config.corriente_max_cierre_aisladero_a=cierre
        usuarios_totales=int(data.get("usuarios_totales_cens",0))
        if usuarios_totales<0:raise ValueError("La cantidad total de usuarios CENS no puede ser negativa.")
        config.usuarios_totales_cens=usuarios_totales
        campos_texto=("maniobra_inicio_desenergizacion","maniobra_fin_desenergizacion","maniobra_inicio_energizacion","maniobra_fin_energizacion")
        for campo in campos_texto:
            valor=str(data.get(campo,"")).strip()
            if not valor:raise ValueError("Los cuatro textos automáticos de maniobra son obligatorios.")
            if len(valor)>1000:raise ValueError("Cada texto automático debe tener máximo 1000 caracteres.")
            setattr(config,campo,valor)
        config.save(update_fields=["corriente_max_apertura_aisladero_a","corriente_max_cierre_aisladero_a","usuarios_totales_cens",*campos_texto,"actualizado_en"])
        return JsonResponse({"ok":True,**_config_maniobras_json(config)})
    except (TypeError,ValueError) as exc:
        return JsonResponse({"error":str(exc)},status=400)
